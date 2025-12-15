import os
import random
import string
from datetime import datetime
from django.conf import settings
from openpyxl import load_workbook
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile


def procesarSubidaArchivo(archivo):
    carpetaRelativa, _ = crearCarpeta()
    nombreUnico = generarNombreUnico(archivo)

    rutaArchivo = f"{carpetaRelativa}/{nombreUnico}"

    guardarArchivo(rutaArchivo, archivo)

    return f"{rutaArchivo}"
  
def crearCarpeta():
    fecha = datetime.now().strftime("%Y-%m-%d")
    carpetaRelativa = f"FileStore/{fecha}"
    return carpetaRelativa, None

def generarNombreUnico(archivo):
    extension = os.path.splitext(archivo.name)[1]
    caracteres = string.ascii_letters + string.digits
    nombre = ''.join(random.choices(caracteres, k=10))    
    return nombre + extension

def guardarArchivo(rutaArchivo,archivo):
    default_storage.save(
        rutaArchivo,
        ContentFile(archivo.read())
    )
  
def obtenerFilasDeExcel(archivo, ignorarColumnas, extraColunmnas):
    wb = load_workbook(archivo)
    ws = wb.active

    headers = [c.value for c in ws[1]]

    fechaActual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    edicionValor = f"{fechaActual}, Sistema , Dato ingreso por lote"

    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        filaDict = {}

        for i, valor in enumerate(row):
            col = headers[i]

            if ignorarColumnas and col in ignorarColumnas:
                continue

            filaDict[col] = valor

        filaDict['ediciones'] = edicionValor

        if extraColunmnas:
            filaDict.update(extraColunmnas)

        filas.append(filaDict)

    return filas

